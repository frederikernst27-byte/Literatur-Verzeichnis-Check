"""Schreibt die Prüfergebnisse als Excel-Tabelle (1 oder 2 Sheets)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .classify import STATUS_MINOR_ISSUES, STATUS_NOT_FOUND, STATUS_OK, STATUS_UNCLEAR, Result

STATUS_COLORS = {
    STATUS_OK: "C6EFCE",
    STATUS_MINOR_ISSUES: "FFEB9C",
    STATUS_NOT_FOUND: "FFC7CE",
    STATUS_UNCLEAR: "D9D9D9",
}

HEADERS = [
    "Nr.", "Original-Zitat", "Status", "Gefundene Quelle",
    "Abweichungen", "Prüfmethode", "Konfidenz (%)", "Link",
]

COL_WIDTHS = [6, 50, 28, 45, 45, 14, 12, 14]


def _write_sheet(ws, results: list[Result], title: str) -> None:
    ws.title = title
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        ws.append([
            r.number,
            r.original_citation,
            r.status,
            r.found_source or "",
            "; ".join(r.discrepancies),
            r.method,
            round(r.confidence, 1),
            "",
        ])
        row = ws.max_row
        fill_color = STATUS_COLORS.get(r.status)
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = fill

        if r.url:
            link_cell = ws.cell(row=row, column=len(HEADERS))
            link_cell.value = "Zum Paper"
            link_cell.hyperlink = r.url
            link_cell.font = Font(color="0563C1", underline="single")

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def export_to_excel(results: list[Result], overflow: list[Result], output_path: str) -> None:
    wb = Workbook()
    _write_sheet(wb.active, results, "Literaturpruefung")

    if overflow:
        ws2 = wb.create_sheet("KI-Runde 2")
        _write_sheet(ws2, overflow, "KI-Runde 2")

    wb.save(output_path)
